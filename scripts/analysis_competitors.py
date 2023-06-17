
import os
import re
import csv
import json
import openpyxl
import asyncio
import logging
import timeit
import numpy as np

from scipy.stats import pearsonr
from websockets import connect
from logger import Logger

from web3 import Web3
from web3.middleware import geth_poa_middleware
from configs.config import CONNECTION, NETWORK, BNB48, load_provider
from analysis_utils import convert_tar_to_pattern, line_time_parser, line_json_parser, unix_time_ms, unix_to_readable, unix_to_readable_ms


COMPETIPORS = {
    "BSC": [
        '0xf060C20bcB2e981cA6d9ec1bE6702Bd762D2678F',
        '0x6b18AF935E9a378eEFfD7324b7f37E1e419b4215',
        '0xb6E496A4Aa6353Ebf6dE6f354ADCCaB513d3a566',
        '0x53dcd3c3Ef1DFa16D4c2d8D053dbdD379C70b6d0',
        '0x3a3452921Ed20703d95238F58EA5Ad7eDF0A495F',
        '0xe3c941Ff0cdA0DD6f0AF092039012266BA620Bb6',
        '0xEe85D50607eD2d1413EABf789eCA6f138A9C362F',
        '0xEaD160F602Ccd12908c100263c4FD8a7ED8f979B',
        '0x09844cb359f768a2a7D185dBf2d972dfb019FE53',
        '0xfb002d7615b9782b7cC8A85aEb34161c78952bd4',

        '0xC1BE2a2290DBB51dDF97869BEcFEb758D6c00230',
        '0xBc12B1fFE8C0c8Aa0004ECfB6969b95aAa1727E0',
        '0x2B1a7a457a2c55BA1e03C087CC3e4E5B05b6360F',

        # 0508
        '0xECF32a129124Dc322bDCe86A690D91B9C7b46d78',
        '0x4c3E78594F12973CE594C29c4ac18195E5485381',
        '0xEB8Fea72614b75024b83D6cdE286739DB501125B',

        # 0515
        '0x832Fa33Ac23ff535A959ceFCC5B3a95d3DAe53B0',
        '0xab0Cee728CFF7c57FAb409bb7022e47F2BC46D96',

        # 0605 ADDED
        '0xF4EF6294eB7DF3e6cD6f6B0fBb4eAe977119dF86',
        '0x5e4d1F5BDef71Dc9ACBe07444162896AA7098a32',
        '0xE8871c8aD85cF5C1C018638298838bf3c450b716'
    ],
    "Ethereum": [
        "0x9d5A494Cec2934Dc01Ec1cAF595840450Bd30f9B",
    ]
}

PREFIX = "./" # "/data/fydeng/"
HASH_ZERO = "0x0000000000000000000000000000000000000000000000000000000000000000"

# w3 = Web3(load_provider('http_local'))
w3 = Web3(load_provider('http_ym'))
w3.middleware_onion.inject(geth_poa_middleware, layer=0)


def pending_callback(message, logger):
    logger.info(f'recieve new message: {message}')


def convert_to_txfilter():
    res = []
    for comp in COMPETIPORS[NETWORK]:
        res.append(
            {'to': comp}
        )

    return res


async def get_pending_transactions_light(callback, logger):
    counter = 0
    while True:
        counter += 1
        try:
            async with connect(CONNECTION[NETWORK]['light']['url'], ping_interval=None,
                            extra_headers={'auth': CONNECTION[NETWORK]['light']['auth']}) as ws:
                await ws.send(
                    json.dumps({
                        'm': 'subscribe',
                        'p': 'txpool',
                        'tx_filters': convert_to_txfilter()
                    }))
                subscription_response = await ws.recv()
                logger.info(subscription_response.strip())

                while True:
                    try:
                        message = await asyncio.wait_for(ws.recv(), timeout=3600)
                    except asyncio.TimeoutError:
                        logger.error("ws Timeout error: no signal")
                        continue

                    message = json.loads(message)
                    txs = message['Txs']

                    # one signal at a time
                    for tx in txs:
                        t = tx['Tx']
                        t['from'] = tx['From']
                        callback(t, logger)

        except Exception as e:
            logger.error(f"unable to connect to light node: {e}")

        if counter >= 3:
            raise Exception("connect to light nodes too many times")
        else:
            logger.info("try to reconnect to light node")


def read_and_parse_from_folder(path, match_file, targets_line, log_filter):
    res = {}
    files = os.listdir(path)
    files.sort()
    for file in files:
        f = path + file
        if os.path.isdir(f):
            continue

        if not match_file in file:
            continue

        print("read file ", f)
        res.update(read_and_parse(f, targets_line, log_filter))

    return res


def read_and_parse(file, targets, log_filter):
    f = open(file)
    iter_f = iter(f)

    file_date = file.split(".log")[0].split("_")[-1]
    patterns = convert_tar_to_pattern(targets)

    wb = openpyxl.Workbook()
    res = {}
    for line in iter_f:
        if line.find(targets[0]) > -1:
            print(line)
            time_onchain = line_time_parser(line)
            tx_hash, block_num, tx_index, borrower, debt, revenue = log_filter(line, patterns[0])
            [time_sig, time_send], sig_hash, status = read_and_parse_file1(file_date, borrower, unix_time_ms(time_onchain))
            
            block = w3.eth.get_block(block_num)
            if tx_hash == HASH_ZERO:
                try:
                    tx_hash = block['transactions'][tx_index].hex()
                except:
                    print(f"error: blocknum {block_num} txs {len(block['transactions'])} index {tx_index}")
                    continue

            validator = block['miner']
            if validator in BNB48:
                is_bnb48 = True
            else:
                is_bnb48 = False

            tx = w3.eth.get_transaction(tx_hash)
            contract = tx['to']
            if contract not in COMPETIPORS[NETWORK]:
                is_comp = False
                print(f"not in listening: {contract}")
            else:
                is_comp = True

            time_competitor = read_and_parse_file2(file_date, tx_hash)
            if time_competitor == 0:
                is_private_send = True 
            else:
                is_private_send = False

            corrcoef = 0
            if time_sig != 0: # and time_competitor != 0:
                if time_competitor == 0:
                    corrcoef = read_and_parse_file2_ext(file_date, tx_hash, sig_hash, time_sig, [time_sig, borrower[2:], debt[2:]], wb, True)
                else:
                    corrcoef = read_and_parse_file2_ext(file_date, tx_hash, sig_hash, time_sig, [time_competitor, borrower[2:], debt[2:]], wb, False)

            print("=================================================")

        # key = str(block_num) + ":" + str(tx_index).zfill(3)
        key = tx_hash
        res[key] = [tx_hash, contract, borrower, sig_hash, time_sig, time_send, time_competitor, unix_time_ms(time_onchain), time_onchain, corrcoef, revenue, is_comp, validator, status, is_bnb48, is_private_send]
    
    wb.save("./outputs/competitions" + file_date + "_details.xlsx")
    return res


def read_and_parse_file1(file_date, target, default_time):
    file = PREFIX + "liquidation_bsc_compound_venus_" + file_date + ".log"
    f = open(file)
    iter_f = iter(f)

    log_time = 0
    log_time2 = 0
    switch = False
    switch_sig = False
    sig_hash = ""
    status = ""
    for line in iter_f:
        if line.find("new message received: ") > -1:
            temp = line
            switch_sig = True

        if switch_sig and line.find(target) > -1 and line.find("liquidation") > -1:
            log_time = unix_time_ms(line_time_parser(temp))
            if log_time + 60 < default_time:
                log_time = 0
                continue

            switch_sig = False
            switch = True
            sig_hash = logs_parse_sig(temp)
            print(temp)
            print(line) 

        if switch and (line.find("bnb48 send") > -1 or line.find("hf liquidation finish: ") > -1):
            log_time2 = unix_time_ms(line_time_parser(line))
            print(line)

            if line.find("bnb48 send success: ") > -1:
                gas_price = logs_parse_bnb48(line)
                status = str(gas_price)
            elif line.find("bnb48 send error: ") > -1:
                status = "error"
            elif line.find("bnb48 send fails: ") > -1:
                status = "fails"
            elif line.find("bnb48 send skipped: ") > -1:
                status = "skipped" 
            else:
                status = "unknown"
            
            switch = False
            break

    return [log_time, log_time2], sig_hash, status


p_bnb48 = re.compile('.*' + "bnb48 send success: " + '(.*)', re.S)
def logs_parse_bnb48(line):
    res_js = line_json_parser(line, p_bnb48)
    gas_price = res_js.get('gas_price')
    return gas_price 


p_sig = re.compile('.*' + "new message received: " + '(.*)', re.S)
def logs_parse_sig(line):
    res_js = line_json_parser(line, p_sig)
    hash = res_js.get('hash')
    return hash


def read_and_parse_file2(file_date, target):
    file = PREFIX + "liquidations_competitor_pendings_" + file_date + ".log"
    f = open(file)
    iter_f = iter(f)

    log_time = 0
    for line in iter_f:
        if line.find(target) > -1:
            log_time = unix_time_ms(line_time_parser(line))
            print(line)
            break

    return log_time


def get_recpt_data(tx_hash):
    try:
        tx_recpt = w3.eth.get_transaction_receipt(tx_hash)
        status = tx_recpt['status']
        block_num = tx_recpt['blockNumber']
        block_index = tx_recpt['transactionIndex']
        position = str(block_num) + ":" + str(block_index).zfill(3) 
        gas_used = tx_recpt['cumulativeGasUsed']
        gas_price = tx_recpt['effectiveGasPrice']
    except Exception:
        status = ""
        position = ""
        gas_used = ""
        gas_price = ""

    return [tx_hash, position, status, gas_used, gas_price]


def get_tx_data(tx_hash):
    tx = w3.eth.get_transaction(tx_hash)
    addr = tx['from']
    contract = tx['to']
    nonce = tx['nonce']
    data = tx['input']

    return contract, [addr, nonce, data]


def read_and_parse_file2_ext(file_date, liq_hash, sig_hash, time_sig, targets, wb, switch):
    file = PREFIX + "liquidations_competitor_pendings_" + file_date + ".log"
    f = open(file)
    iter_f = iter(f)

    '''
    with open("./outputs/competitions" + file_date + "_detals.csv", 'a', newline='') as csvfile:
        writer = csv.writer(csvfile)

        if csvfile.tell() == 0:
            writer.writerow(['time', 'bot', 'hash', 'position', 'status', 'gasUsed', 'gasPrice', 'data'])

        # Write the data rows
        writer.writerow([log_time, contract, tx_hash, position, status, gas_used, gas_price, data])
    '''
    ws = wb.create_sheet(liq_hash[:8])
    ws.append(['time', 'bot', 'hash', 'position', 'status', 'gasUsed', 'gasPrice', 'from', 'nonce', 'data'])

    res = get_recpt_data(sig_hash)
    ws.append([unix_to_readable_ms(time_sig), "transmitter"] + res + ["", "", ""]) 

    tar0_arr = [unix_to_readable(int(targets[0]-1)), unix_to_readable(int(targets[0])), unix_to_readable(int(targets[0]+1))]

    for line in iter_f:
        if (line.find(tar0_arr[0]) > -1 or line.find(tar0_arr[1]) > -1 or line.find(tar0_arr[2]) > -1) and (line.find(targets[1]) > -1 or line.find(targets[2]) > -1):
            log_time = line_time_parser(line)
            tx_hash, contract, gas_price, addr_from, nonce, data = logs_parser_comp(line)

            res = get_recpt_data(tx_hash)
            ws.append([log_time, contract] + res + [addr_from, nonce, data])

    res_coef = cal_corrcoef(ws)

    if switch:
        arr1 = get_recpt_data(liq_hash)
        contract, arr2 = get_tx_data(liq_hash) 
        ws.append(["", contract] + arr1 + arr2) 
 
    return res_coef



p_comp = re.compile('.*' + "recieve new message: " + '(.*)', re.S)
def logs_parser_comp(line):
    res_js = line_json_parser(line, p_comp)

    tx_hash = res_js.get('hash', None)
    contract =  res_js.get('to') 
    addr_from = res_js.get('from')
    nonce = int(res_js.get('nonce'), 16)
    gas_price = int(res_js.get('gasPrice'), 16)
    data = res_js.get('input')

    return tx_hash, contract, gas_price, addr_from, nonce, data


def logs_parser_onchain(line, p):
    res_js = line_json_parser(line, p)
    tx_hash = res_js.get('txHash', None)
    block_num = res_js.get('blockNumber') 
    tx_index = res_js.get('index')
    borrower = res_js.get('borrower') 
    debt = res_js.get('debt').lower()

    revenue = res_js.get('revenue')
    if isinstance(revenue, list):
        revenue = revenue[1] - revenue[2]
    revenue = revenue/10**18

    return tx_hash, block_num, tx_index, borrower, debt, revenue


def process():
    date = "20230610"
    res = read_and_parse_from_folder(PREFIX, "liquidations_onchain_" + date, ["liquidationCalls: "], logs_parser_onchain)
    print(res)

    wb = openpyxl.Workbook()
    ws = wb.create_sheet('ALL')
    ws.append(["liq_hash", "bot", "borrower", "sig_hash", "time_sig", "time_send", "time_comp", "time_onchain", "on_chain", "corrcoef", "revenue", "is_comp", "validator", "bnb48_send_status", "is_bnb48", "is_priv_send"])
    for data in res.values():
        ws.append(data)
    wb.save("./outputs/competitions" + date + ".xlsx")

    '''
    with open("./outputs/competitions" + date + ".csv", 'w', newline='') as csvfile:
        csv_writer = csv.writer(csvfile)
        csv_writer.writerow(["signal", "my-send", "competitor", "on_chain", "revenue", "is_comp", "is_BNB48"])
        
        for data in res.values():
            csv_writer.writerow(data)
    '''


def sort_with_index(arr):
    sorted_arr = sorted(arr)
    index_arr = [sorted_arr.index(x) for x in arr]
    return sorted_arr, index_arr


def needleman_wunsch(seq1, seq2, match_score=1, mismatch_score=-1, gap_score=-1):
    """
    Perform global sequence alignment using the Needleman-Wunsch algorithm.
    """
    n = len(seq1)
    m = len(seq2)

    # Initialize the score matrix
    score_matrix = [[0] * (m + 1) for i in range(n + 1)]

    # Initialize the traceback matrix
    traceback_matrix = [[0] * (m + 1) for i in range(n + 1)]

    # Initialize the first row and column of the score matrix
    for i in range(n + 1):
        score_matrix[i][0] = gap_score * i
        traceback_matrix[i][0] = "U"

    for j in range(m + 1):
        score_matrix[0][j] = gap_score * j
        traceback_matrix[0][j] = "L"

    # Fill in the rest of the score matrix
    for i in range(1, n + 1):
        for j in range(1, m + 1):
            if seq1[i-1] == seq2[j-1]:
                match = match_score
            else:
                match = mismatch_score

            diag_score = score_matrix[i-1][j-1] + match
            up_score = score_matrix[i-1][j] + gap_score
            left_score = score_matrix[i][j-1] + gap_score

            max_score = max(diag_score, up_score, left_score)

            if max_score == diag_score:
                traceback_matrix[i][j] = "D"
            elif max_score == up_score:
                traceback_matrix[i][j] = "U"
            else:
                traceback_matrix[i][j] = "L"

            score_matrix[i][j] = max_score

    # Traceback to find the optimal alignment
    align1 = ""
    align2 = ""
    i = n
    j = m

    while i > 0 or j > 0:
        if traceback_matrix[i][j] == "D":
            align1 = seq1[i-1] + align1
            align2 = seq2[j-1] + align2
            i -= 1
            j -= 1
        elif traceback_matrix[i][j] == "U":
            align1 = seq1[i-1] + align1
            align2 = "-" + align2
            i -= 1
        else:
            align1 = "-" + align1
            align2 = seq2[j-1] + align2
            j -= 1

    return align1, align2, score_matrix[n][m]



def trim_none(a, b):
    a_trimmed = [elem for elem in a if elem != '']
    b_trimmed = [b[i] for i, elem in enumerate(a) if elem != '']

    return a_trimmed, b_trimmed 


def analysis():
    workbook = openpyxl.load_workbook('./outputs/competitions20230506_details.xlsx')
    worksheet = workbook['0xe52e97']
    cal_corrcoef(worksheet)  # 0xe52e97: 0.2606060606060606


def cal_corrcoef(worksheet):
    column_a = [cell.value for cell in worksheet['A']][1:]
    column_d = [cell.value for cell in worksheet['D']][1:]

    arr2, arr1 = trim_none(column_d, column_a)
    if len(arr1) <= 1:
        return 0

    _, seq1 = sort_with_index(arr1)
    _, seq2 = sort_with_index(arr2)

    # res = needleman_wunsch(seq1, seq2)
    res = np.corrcoef(seq1, seq2)
    comp = pearsonr(seq1, seq2)[0]
    # print(res[0][1], comp)

    return res[0][1]


def exp():
    logger = Logger(log_file_name="liquidations_competitor_pendings", log_level=logging.DEBUG, logger_name="compet").get_log()
    asyncio.run(get_pending_transactions_light(pending_callback, logger))


if __name__ == '__main__':
    process()
    # exp()
    # analysis()
